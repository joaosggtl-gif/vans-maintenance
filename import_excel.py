#!/usr/bin/env python3
"""
Excel to Fleet Dashboard Importer
Converts the Excel maintenance file to JSON format for the dashboard
"""

import json
import re
from datetime import datetime
import openpyxl

def generate_id():
    """Generate a unique ID"""
    import random
    import time
    return f"id_{int(time.time())}_{random.randint(1000, 9999)}"

def clean_reg(reg):
    """Clean and normalize vehicle registration"""
    if not reg:
        return None
    reg = str(reg).strip().upper()
    reg = re.sub(r'\s+', '', reg)  # Remove all spaces
    return reg if reg and reg != 'NONE' else None

def parse_date(date_val):
    """Parse date from various formats"""
    if not date_val:
        return None
    if isinstance(date_val, datetime):
        return date_val.strftime('%Y-%m-%d')
    try:
        return datetime.strptime(str(date_val), '%Y-%m-%d').strftime('%Y-%m-%d')
    except:
        return None

def safe_str(val):
    """Safely convert value to string"""
    if val is None:
        return ''
    return str(val).strip()

def extract_vehicles_from_van_list(sheet):
    """Extract vehicle information from Van list sheet"""
    vehicles = {}

    for row in sheet.iter_rows(min_row=3, values_only=True):
        if not row or not row[2]:  # REG NUMBER is column C (index 2)
            continue

        reg = clean_reg(row[2])
        if not reg:
            continue

        provider = safe_str(row[0]) if row[0] else 'Unknown'
        brand = safe_str(row[1]) if row[1] else 'Ford'

        if reg not in vehicles:
            vehicles[reg] = {
                'id': generate_id(),
                'reg': reg,
                'make': brand,
                'model': 'Transit' if 'Ford' in brand else 'Sprinter' if 'Merc' in brand else '',
                'year': '2021',  # Most are 71 plates
                'initialMileage': 0,
                'color': 'White',
                'notes': f'Provider: {provider}',
                'services': [],
                'createdAt': datetime.now().isoformat()
            }

    return vehicles

def extract_services_from_yearly_sheet(sheet, year_name):
    """Extract services from yearly sheets (2022, 2023, 2024, 2025)"""
    services = []

    # Find header row
    header_row = None
    for i, row in enumerate(sheet.iter_rows(min_row=1, max_row=5, values_only=True)):
        if row and any('Date' in str(c) if c else False for c in row):
            header_row = i + 1
            break

    if not header_row:
        header_row = 1

    for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
        if not row:
            continue

        # Different column structures for different years
        date_val = None
        reg = None
        location = ''
        work = ''
        make_model = ''

        # Try to identify columns
        for i, val in enumerate(row):
            if isinstance(val, datetime):
                date_val = val
            elif val and isinstance(val, str):
                val_upper = val.strip().upper()
                # Check if it looks like a registration
                if re.match(r'^[A-Z]{2}\d{2}\s?[A-Z]{3}$', val_upper.replace(' ', '')):
                    reg = clean_reg(val)
                elif 'FORD' in val_upper or 'TRANSIT' in val_upper or 'MERCEDES' in val_upper:
                    make_model = val
                elif 'REDRUTH' in val_upper or 'SITE' in val_upper or 'GARAGE' in val_upper or 'FORD' in val_upper:
                    location = val

        # Try positional extraction based on common patterns
        if not reg and len(row) > 2:
            # Year sheets typically have: index, date, reg, lease, make, location, work
            for j, val in enumerate(row[:6]):
                if val and isinstance(val, str) and re.match(r'^[A-Z]{2}\d{2}\s?[A-Z]{3}$', str(val).strip().upper().replace(' ', '')):
                    reg = clean_reg(val)
                    break

        if not date_val:
            # Find first datetime in row
            for val in row:
                if isinstance(val, datetime):
                    date_val = val
                    break

        if not reg:
            continue

        # Extract work description - usually one of the last meaningful columns
        for val in reversed(row):
            if val and isinstance(val, str) and len(val) > 3 and val.lower() not in ['yes', 'no', 'ok', 'none']:
                if not any(x in val.lower() for x in ['ford', 'transit', 'redruth', 'arval', 'leaseplan']):
                    work = val
                    break

        if not work:
            work = 'Maintenance service'

        services.append({
            'reg': reg,
            'date': parse_date(date_val) if date_val else f'{year_name}-01-01',
            'work': work,
            'location': location or 'Redruth',
            'provider': 'In-house',
            'mileage': 0
        })

    return services

def extract_services_from_vehicle_sheet(sheet, reg):
    """Extract services from individual vehicle sheets"""
    services = []

    # Find header row (usually row 4)
    header_row = 4

    for row in sheet.iter_rows(min_row=5, values_only=True):
        if not row or not row[0]:
            continue

        date_val = row[0]
        if not isinstance(date_val, datetime):
            # Check if this is a second header row (2025 format)
            if row[0] and 'Date' in str(row[0]):
                continue
            continue

        # Old format: Date, Invoice N, Service, Description, Miles, Price, Notes
        # New format: Date, Reg Number, Provider, Make/Model, Garage, Work carry out, Details, Mileage

        invoice = safe_str(row[1]) if len(row) > 1 else ''
        service_type = safe_str(row[2]) if len(row) > 2 else ''
        description = safe_str(row[3]) if len(row) > 3 else ''
        miles = row[4] if len(row) > 4 and isinstance(row[4], (int, float)) else 0

        # Determine provider
        provider = 'In-house'
        if invoice:
            if 'Peter' in invoice or 'Adam' in invoice or 'Petter' in invoice:
                provider = invoice
            elif 'LEASEPLAN' in invoice.upper():
                provider = 'Leaseplan'
            elif 'FORD' in invoice.upper() or 'Trust' in invoice.lower():
                provider = 'Trust Ford'

        # Determine location
        location = 'On Site'
        if service_type:
            if 'MJB' in service_type or 'tyres' in service_type.lower():
                location = 'MJB Tyres'
            elif 'Ford' in service_type:
                location = 'Ford Dealer'
            elif 'Motor parts' in service_type:
                location = 'Motor Parts Direct'

        work = description if description else service_type if service_type else 'Service'

        services.append({
            'id': generate_id(),
            'date': parse_date(date_val),
            'work': work,
            'duration': '',
            'provider': provider,
            'mileage': int(miles) if miles else 0,
            'location': location,
            'cost': '',
            'notes': f'Invoice: {invoice}' if invoice and invoice != provider else '',
            'createdAt': datetime.now().isoformat()
        })

    return services

def main():
    print("=" * 60)
    print("Fleet Maintenance Excel Importer")
    print("=" * 60)

    xlsx_file = "General vans maintenance - Updated.xlsx"

    print(f"\nLoading: {xlsx_file}")
    wb = openpyxl.load_workbook(xlsx_file)

    print(f"Found {len(wb.sheetnames)} sheets:")
    for name in wb.sheetnames:
        print(f"  - {name}")

    # Initialize vehicles dictionary
    vehicles = {}

    # 1. Extract vehicles from "Van list" sheet
    if 'Van list ' in wb.sheetnames:
        print("\n[1/4] Extracting vehicles from 'Van list'...")
        vehicles = extract_vehicles_from_van_list(wb['Van list '])
        print(f"  Found {len(vehicles)} vehicles")

    # 2. Process individual vehicle sheets (they have the most detailed history)
    print("\n[2/4] Processing individual vehicle sheets...")
    vehicle_sheets = [name for name in wb.sheetnames if re.match(r'^[A-Z]{2}\d{2}[A-Z]{3}$', name.replace(' ', ''))]

    for sheet_name in vehicle_sheets:
        reg = clean_reg(sheet_name)
        print(f"  Processing: {reg}")

        if reg not in vehicles:
            vehicles[reg] = {
                'id': generate_id(),
                'reg': reg,
                'make': 'Ford',
                'model': 'Transit',
                'year': '2021',
                'initialMileage': 0,
                'color': 'White',
                'notes': '',
                'services': [],
                'createdAt': datetime.now().isoformat()
            }

        sheet_services = extract_services_from_vehicle_sheet(wb[sheet_name], reg)
        vehicles[reg]['services'].extend(sheet_services)
        print(f"    Added {len(sheet_services)} services")

    # 3. Process yearly sheets for any missing services
    print("\n[3/4] Processing yearly sheets...")
    yearly_sheets = ['2022', '2023 ', '2024', '2025']

    for year_name in yearly_sheets:
        if year_name in wb.sheetnames:
            year_clean = year_name.strip()
            print(f"  Processing: {year_clean}")
            year_services = extract_services_from_yearly_sheet(wb[year_name], year_clean)

            # Add services to vehicles (avoid duplicates by checking date+work)
            added = 0
            for svc in year_services:
                reg = svc['reg']
                if reg not in vehicles:
                    vehicles[reg] = {
                        'id': generate_id(),
                        'reg': reg,
                        'make': 'Ford',
                        'model': 'Transit',
                        'year': '2021',
                        'initialMileage': 0,
                        'color': 'White',
                        'notes': '',
                        'services': [],
                        'createdAt': datetime.now().isoformat()
                    }

                # Check for duplicates
                existing = vehicles[reg]['services']
                is_duplicate = any(
                    s.get('date') == svc['date'] and s.get('work', '').lower() == svc['work'].lower()
                    for s in existing
                )

                if not is_duplicate:
                    vehicles[reg]['services'].append({
                        'id': generate_id(),
                        'date': svc['date'],
                        'work': svc['work'],
                        'duration': '',
                        'provider': svc['provider'],
                        'mileage': svc.get('mileage', 0),
                        'location': svc['location'],
                        'cost': '',
                        'notes': '',
                        'createdAt': datetime.now().isoformat()
                    })
                    added += 1

            print(f"    Added {added} new services")

    # 4. Update mileage from services
    print("\n[4/4] Updating mileage data...")
    for reg, vehicle in vehicles.items():
        max_mileage = 0
        for svc in vehicle['services']:
            if svc.get('mileage', 0) > max_mileage:
                max_mileage = svc['mileage']
        if max_mileage > 0:
            vehicle['initialMileage'] = max_mileage

    # Create output structure
    output = {
        'vehicles': list(vehicles.values()),
        'lastUpdated': datetime.now().isoformat()
    }

    # Calculate statistics
    total_services = sum(len(v['services']) for v in vehicles.values())

    print("\n" + "=" * 60)
    print("IMPORT SUMMARY")
    print("=" * 60)
    print(f"Total Vehicles: {len(vehicles)}")
    print(f"Total Services: {total_services}")
    print("\nVehicles imported:")
    for reg, vehicle in sorted(vehicles.items()):
        print(f"  {reg}: {len(vehicle['services'])} services")

    # Save to JSON
    output_file = "fleet_import_data.json"
    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(output, f, indent=2, ensure_ascii=False)

    print(f"\n✓ Data saved to: {output_file}")
    print("\nTo import into the dashboard:")
    print("  1. Open the dashboard (index.html)")
    print("  2. Click 'Import Data' in the sidebar")
    print("  3. Select 'fleet_import_data.json'")
    print("\n" + "=" * 60)

if __name__ == '__main__':
    main()
